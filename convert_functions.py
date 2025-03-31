from fpdf import FPDF
from openpyxl import Workbook

import os

from tkinter import messagebox

def convert_to_pdf(table, tree, columns):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    if len(columns) <= 5:
        font_size = 12
    elif len(columns) <= 10:
        font_size = 10
    else:
        font_size = 8
    pdf.set_font("Arial", size=font_size)
    pdf.cell(200, 10, f"Table: {table}", 0, 1, "C")
    pdf.ln(10)

    # Exclude specific columns for the items table
    if table == 'items':
        columns_to_include = [col for col in columns if col not in ['Serial Number', 'last_updated']]
    else:
        columns_to_include = columns

    col_widths = [pdf.get_string_width(col) + 10 for col in columns_to_include]
    displayed_rows = [tree.item(item)['values'] for item in tree.get_children()]
    for row in displayed_rows:
        for i, item in enumerate(row):
            if columns[i] in columns_to_include:
                col_widths[columns_to_include.index(columns[i])] = max(col_widths[columns_to_include.index(columns[i])], pdf.get_string_width(str(item)) + 10)

    # Print the column headers
    for i, col in enumerate(columns_to_include):
        pdf.cell(col_widths[i], 10, col, 1, 0, "C")
    pdf.ln()

    # Print the rows
    for row in displayed_rows:
        row_height = 20 if table == 'items' else 10  # Adjust row height for items table to accommodate QR code
        max_y = pdf.get_y() + row_height
        for i, item in enumerate(row):
            if columns[i] in columns_to_include:
                pdf.cell(col_widths[columns_to_include.index(columns[i])], row_height, str(item), border=1, align='C')
        if table == 'items':
            # Reuse the saved QR code for the row
            qr_file = f"qr_{row[0]}.png"
            x = pdf.get_x() + (col_widths[-1] - 20) / 2  # Center the QR code in the column
            y = max_y - row_height + (row_height - 20) / 2  # Center the QR code vertically
            pdf.image(qr_file, x=x, y=y, w=20, h=20)  # Resize the QR code to be smaller in the PDF
        pdf.set_y(max_y)

    if not os.path.exists('pdf'):
        os.makedirs('pdf')

    pdf_folder = os.path.join('pdf', table)
    if not os.path.exists(pdf_folder):
        os.makedirs(pdf_folder)

    pdf_file = os.path.join(pdf_folder, f"{table}.pdf")
    counter = 1
    while os.path.exists(pdf_file):
        pdf_file = os.path.join(pdf_folder, f"{table}_{counter}.pdf")
        counter += 1
    pdf.output(pdf_file)
    messagebox.showinfo("Success", f"Table {table} has been converted to PDF")
    os.startfile(pdf_file)
    return

def convert_to_excel(table, rows, columns):
    wb = Workbook()
    ws = wb.active
    ws.title = table

    # columnHeaders
    for col_num, column_title in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # rows
    for row_num, row_data in enumerate(rows, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    excel_file = f"{table}.xlsx"
    counter = 1
    while os.path.exists(excel_file):
        excel_file = f"{table}_{counter}.xlsx"
        counter += 1
    wb.save(excel_file)
    messagebox.showinfo("Success", f"Table {table} has been converted to Excel")
    os.startfile(excel_file)
    return

def generate_qr_pdf(table, tree):
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Arial", size=8)

        # Create the folder structure if it doesn't exist
        qr_folder = os.path.join("pdf", table, "qr_codes")
        if not os.path.exists(qr_folder):
            os.makedirs(qr_folder)

        # Page width and QR code dimensions
        page_width = 190  # A4 page width in mm (excluding margins)
        qr_width = 25     # Width of each QR code
        qr_height = 25    # Height of each QR code
        spacing = 10      # Spacing between QR codes
        max_per_row = page_width // (qr_width + spacing)  # Maximum QR codes per row

        x_start = 10  # Starting x position (left margin)
        y_start = 20  # Starting y position (top margin)
        x = x_start
        y = y_start

        # Iterate through the rows and add QR codes to the PDF
        for row in tree.get_children():
            values = tree.item(row, "values")
            if table == "items":
                qr_file = f"qr_{values[0]}.png"
                if os.path.exists(qr_file):
                    # Add the QR code image
                    pdf.image(qr_file, x=x, y=y, w=qr_width, h=qr_height)

                    # Add the caption below the QR code
                    pdf.set_xy(x, y + qr_height)
                    pdf.cell(qr_width, 5, f"{values[0]}", 0, 0, "C")

                    # Move to the next position
                    x += qr_width + spacing
                    if x + qr_width > page_width:  # If the next QR code exceeds the page width
                        x = x_start  # Reset x to the starting position
                        y += qr_height + 15  # Move to the next row

                    # Check if the next row exceeds the page height
                    if y + qr_height + 15 > 277:  # A4 page height in mm (excluding margins)
                        pdf.add_page()
                        x = x_start
                        y = y_start

        # Save the PDF
        qr_pdf_file = os.path.join(qr_folder, f"{table}_qr_codes.pdf")
        counter = 1
        while os.path.exists(qr_pdf_file):
            qr_pdf_file = os.path.join(qr_folder, f"{table}_qr_codes{counter}.pdf")
            counter += 1
        pdf.output(qr_pdf_file)
        messagebox.showinfo("Success", f"QR Codes for table {table} have been saved to {qr_pdf_file}")
        os.startfile(qr_pdf_file)

if __name__ == '__main__':
    messagebox.showwarning('Warning', 'Wrong way to run the module. Run the index.py file instead.')